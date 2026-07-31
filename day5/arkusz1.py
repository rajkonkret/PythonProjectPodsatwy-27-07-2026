import openpyxl
from openpyxl.worksheet import worksheet

# pip install openpyxl

# działania na plikch excel, formatowanie, stylizacja, wykresy

workbook = openpyxl.load_workbook('sales.xlsx')
worksheet = workbook.active

print(worksheet)  # <Worksheet "Arkusz1">

print(worksheet.title)  # Arkusz1

lista = []
for i in worksheet:
    print(i)
# (<Cell 'Arkusz1'.A1>, <Cell 'Arkusz1'.B1>, <Cell 'Arkusz1'.C1>)
# (<Cell 'Arkusz1'.A2>, <Cell 'Arkusz1'.B2>, <Cell 'Arkusz1'.C2>)
# (<Cell 'Arkusz1'.A3>, <Cell 'Arkusz1'.B3>, <Cell 'Arkusz1'.C3>)
# (<Cell 'Arkusz1'.A4>, <Cell 'Arkusz1'.B4>, <Cell 'Arkusz1'.C4>)
# (<Cell 'Arkusz1'.A5>, <Cell 'Arkusz1'.B5>, <Cell 'Arkusz1'.C5>)

for i in range(worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        lista.append(col[i].value)

print(lista)
